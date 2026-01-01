"""
RTA Break Tracker - Web Application
Flask-based web app for tracking agent breaks
"""
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, Response
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
from pathlib import Path
import bcrypt
import os
import uuid
import io

from config import (
    SECRET_KEY, SQLALCHEMY_DATABASE_URI, UPLOAD_FOLDER, 
    ALLOWED_EXTENSIONS, BREAK_DURATIONS, BREAK_INFO,
    ROLE_AGENT, ROLE_RTM, DEFAULT_USERS, DEBUG, ENV, TIMEZONE
)
import pytz

# Break types that count as working time (meetings/coaching/overtime)
WORKING_TIME_BREAKS = ['coaching_aya', 'coaching_mostafa', 'meeting_team_leader', 'overtime']
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def get_local_time():
    """Get current time in configured timezone"""
    return datetime.now(TIMEZONE)

def to_local_time(dt):
    """Return datetime as-is (already stored in local time)"""
    if dt is None:
        return None
    # Times are stored in local timezone already, just return as-is
    return dt

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = SECRET_KEY
app.config['SQLALCHEMY_DATABASE_URI'] = SQLALCHEMY_DATABASE_URI
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = str(UPLOAD_FOLDER)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# Initialize extensions
db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'

# ==================== MODELS ====================

class User(UserMixin, db.Model):
    """User model for agents and RTM"""
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    full_name = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False, default=ROLE_AGENT)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    breaks = db.relationship('BreakRecord', backref='agent', lazy=True)
    
    def is_rtm(self):
        return self.role == ROLE_RTM
    
    def is_agent(self):
        return self.role == ROLE_AGENT
    
    def set_password(self, password):
        self.password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    def check_password(self, password):
        return bcrypt.checkpw(password.encode('utf-8'), self.password_hash.encode('utf-8'))


class Shift(db.Model):
    """Shift model for tracking agent work schedules"""
    id = db.Column(db.Integer, primary_key=True)
    agent_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    start_time = db.Column(db.Time, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    end_time = db.Column(db.Time, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    
    agent = db.relationship('User', foreign_keys=[agent_id], backref='shifts')
    
    def get_duration_hours(self):
        """Calculate shift duration in hours"""
        start = datetime.combine(self.start_date, self.start_time)
        end = datetime.combine(self.end_date, self.end_time)
        return (end - start).total_seconds() / 3600
    
    def get_shift_date(self):
        """Get the primary shift date (start date) for backward compatibility"""
        return self.start_date
    
    def to_dict(self):
        return {
            'id': self.id,
            'agent_id': self.agent_id,
            'agent_name': self.agent.full_name if self.agent else 'Unknown',
            'start_date': self.start_date.isoformat(),
            'start_time': self.start_time.strftime('%H:%M'),
            'end_date': self.end_date.isoformat(),
            'end_time': self.end_time.strftime('%H:%M'),
            'shift_date': self.start_date.isoformat(),  # For backward compatibility
            'duration_hours': round(self.get_duration_hours(), 2)
        }


class OffDay(db.Model):
    """Off day model for tracking agent days off"""
    id = db.Column(db.Integer, primary_key=True)
    agent_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    off_date = db.Column(db.Date, nullable=False)
    reason = db.Column(db.String(255), default='')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    
    agent = db.relationship('User', foreign_keys=[agent_id], backref='off_days')
    
    def to_dict(self):
        return {
            'id': self.id,
            'agent_id': self.agent_id,
            'agent_name': self.agent.full_name if self.agent else 'Unknown',
            'off_date': self.off_date.isoformat(),
            'reason': self.reason,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


class BreakRecord(db.Model):
    """Break record model"""
    id = db.Column(db.Integer, primary_key=True)
    agent_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    break_type = db.Column(db.String(20), nullable=False)
    start_time = db.Column(db.DateTime, nullable=False)
    start_screenshot = db.Column(db.String(255))
    end_time = db.Column(db.DateTime)
    end_screenshot = db.Column(db.String(255))
    duration_minutes = db.Column(db.Integer)
    is_overdue = db.Column(db.Boolean, default=False)
    notes = db.Column(db.Text, default='')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def is_active(self):
        return self.end_time is None
    
    def get_elapsed_minutes(self):
        if self.start_time:
            # Both times are in local timezone
            now = get_local_time().replace(tzinfo=None)
            elapsed = now - self.start_time
            return max(0, int(elapsed.total_seconds() / 60))
        return 0
    
    def get_break_info(self):
        return BREAK_INFO.get(self.break_type, {"name": self.break_type, "emoji": "⏱️", "color": "#666"})
    
    def get_allowed_duration(self):
        return BREAK_DURATIONS.get(self.break_type, 15)
    
    def is_working_time_break(self):
        """Check if this break counts as working time (coaching/meetings)"""
        return self.break_type in WORKING_TIME_BREAKS
    
    def get_effective_overdue_status(self):
        """Get overdue status, but always False for working time breaks"""
        if self.is_working_time_break():
            return False
        return self.is_overdue
    
    def get_local_start_time(self):
        """Get start time in local timezone"""
        if self.start_time:
            return to_local_time(self.start_time)
        return None
    
    def get_local_end_time(self):
        """Get end time in local timezone"""
        if self.end_time:
            return to_local_time(self.end_time)
        return None
    
    def to_dict(self):
        info = self.get_break_info()
        local_start = self.get_local_start_time()
        local_end = self.get_local_end_time()
        return {
            'id': self.id,
            'agent_id': self.agent_id,
            'agent_name': self.agent.full_name if self.agent else 'Unknown',
            'break_type': self.break_type,
            'break_name': info['name'],
            'break_emoji': info['emoji'],
            'break_color': info['color'],
            'start_time': local_start.isoformat() if local_start else None,
            'end_time': local_end.isoformat() if local_end else None,
            'start_screenshot': self.start_screenshot,
            'end_screenshot': self.end_screenshot,
            'duration_minutes': self.duration_minutes,
            'elapsed_minutes': self.get_elapsed_minutes() if self.is_active() else self.duration_minutes,
            'is_active': self.is_active(),
            'is_overdue': self.get_effective_overdue_status(),  # Use effective status (excludes working time breaks)
            'notes': self.notes or '',
            'allowed_duration': self.get_allowed_duration()
        }


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ==================== HELPERS ====================

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def save_screenshot(file):
    """Save uploaded screenshot and return filename"""
    if file and allowed_file(file.filename):
        # Generate unique filename
        ext = file.filename.rsplit('.', 1)[1].lower()
        filename = f"{uuid.uuid4().hex}.{ext}"
        
        # Create date-based folder
        today = get_local_time().strftime("%Y-%m-%d")
        folder = Path(app.config['UPLOAD_FOLDER']) / today
        folder.mkdir(parents=True, exist_ok=True)
        
        # Save file
        filepath = folder / filename
        file.save(str(filepath))
        
        return f"{today}/{filename}"
    return None


def init_db():
    """Initialize database and create default users"""
    db.create_all()
    
    # Migrate Shift table if needed (add new columns, keep old ones for compatibility)
    try:
        from sqlalchemy import inspect, text
        inspector = inspect(db.engine)
        columns = [col['name'] for col in inspector.get_columns('shift')]
        
        # Check if new columns exist
        has_start_date = 'start_date' in columns
        has_end_date = 'end_date' in columns
        has_old_shift_date = 'shift_date' in columns
        
        # If old column exists but new ones don't, add them
        if has_old_shift_date and (not has_start_date or not has_end_date):
            print("Migrating Shift table: Adding start_date and end_date columns...")
            with db.engine.connect() as conn:
                if not has_start_date:
                    conn.execute(text("ALTER TABLE shift ADD COLUMN start_date DATE"))
                    # Copy data from shift_date to start_date
                    conn.execute(text("UPDATE shift SET start_date = shift_date WHERE start_date IS NULL"))
                if not has_end_date:
                    conn.execute(text("ALTER TABLE shift ADD COLUMN end_date DATE"))
                    # Copy data from shift_date to end_date
                    conn.execute(text("UPDATE shift SET end_date = shift_date WHERE end_date IS NULL"))
                conn.commit()
            print("✅ Shift table migration completed")
    except Exception as e:
        print(f"Note: Could not check/migrate Shift table schema: {e}")
        # Continue anyway - db.create_all() will handle new columns
    
    # Create default users if they don't exist
    for user_data in DEFAULT_USERS:
        if not User.query.filter_by(username=user_data['username']).first():
            user = User(
                username=user_data['username'],
                full_name=user_data['full_name'],
                role=user_data['role']
            )
            user.set_password(user_data['password'])
            db.session.add(user)
    
    db.session.commit()


# ==================== ROUTES ====================

@app.route('/')
def index():
    """Home page - redirect to appropriate view"""
    if current_user.is_authenticated:
        if current_user.is_rtm():
            return redirect(url_for('dashboard'))
        return redirect(url_for('agent_view'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user)
            flash('Logged in successfully!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    """Logout"""
    logout_user()
    flash('Logged out successfully', 'success')
    return redirect(url_for('login'))


@app.route('/agent')
@login_required
def agent_view():
    """Agent view for submitting breaks"""
    if current_user.is_rtm():
        return redirect(url_for('dashboard'))
    
    # Get active break for this agent (exclude punch_in/punch_out as they're auto-completed)
    # Use a more compatible query method
    all_active = BreakRecord.query.filter(
        BreakRecord.agent_id == current_user.id,
        BreakRecord.end_time == None
    ).all()
    active_break = next((b for b in all_active if b.break_type not in ['punch_in', 'punch_out']), None)
    
    # Get today's breaks
    today = get_local_time().date()
    today_breaks = BreakRecord.query.filter(
        BreakRecord.agent_id == current_user.id,
        db.func.date(BreakRecord.start_time) == today
    ).order_by(BreakRecord.start_time.desc()).all()
    
    # Get today's shift for this agent (shift that starts today or includes today)
    today_shift = Shift.query.filter(
        Shift.agent_id == current_user.id,
        Shift.start_date <= today,
        Shift.end_date >= today
    ).first()
    
    # Get upcoming shifts (next 7 days) - shifts that start within next 7 days
    next_week = today + timedelta(days=7)
    upcoming_shifts = Shift.query.filter(
        Shift.agent_id == current_user.id,
        Shift.start_date >= today,
        Shift.start_date <= next_week
    ).order_by(Shift.start_date).all()
    
    # Get upcoming off days (next 7 days)
    upcoming_offdays = OffDay.query.filter(
        OffDay.agent_id == current_user.id,
        OffDay.off_date >= today,
        OffDay.off_date <= next_week
    ).order_by(OffDay.off_date).all()
    
    # Check if today is an off day
    is_off_day = OffDay.query.filter_by(
        agent_id=current_user.id,
        off_date=today
    ).first() is not None
    
    # Check punch in/out status for today
    punch_in = BreakRecord.query.filter(
        BreakRecord.agent_id == current_user.id,
        BreakRecord.break_type == 'punch_in',
        db.func.date(BreakRecord.start_time) == today
    ).first()
    
    punch_out = BreakRecord.query.filter(
        BreakRecord.agent_id == current_user.id,
        BreakRecord.break_type == 'punch_out',
        db.func.date(BreakRecord.start_time) == today
    ).first()
    
    # Determine agent status
    # not_punched_in: hasn't punched in yet today
    # punched_in: punched in but not out
    # punched_out: already punched out for the day
    if not punch_in:
        punch_status = 'not_punched_in'
    elif punch_out:
        punch_status = 'punched_out'
    else:
        punch_status = 'punched_in'
    
    return render_template('agent.html',
        user=current_user,
        active_break=active_break,
        today_breaks=today_breaks,
        today_shift=today_shift,
        upcoming_shifts=upcoming_shifts,
        upcoming_offdays=upcoming_offdays,
        is_off_day=is_off_day,
        punch_status=punch_status,
        punch_in_time=punch_in.start_time if punch_in else None,
        punch_out_time=punch_out.start_time if punch_out else None,
        break_types=BREAK_INFO,
        break_durations=BREAK_DURATIONS
    )


@app.route('/dashboard')
@login_required
def dashboard():
    """RTM Dashboard"""
    if not current_user.is_rtm():
        return redirect(url_for('agent_view'))
    
    # Get filter parameters
    date_filter = request.args.get('date', get_local_time().strftime('%Y-%m-%d'))
    agent_filter = request.args.get('agent', '')
    type_filter = request.args.get('type', '')
    
    # Get all agents
    agents = User.query.filter_by(role=ROLE_AGENT).order_by(User.full_name).all()
    
    # Get stats (exclude punch_in/punch_out as they're attendance records, not breaks)
    today = get_local_time().date()
    total_breaks_today = BreakRecord.query.filter(
        db.func.date(BreakRecord.start_time) == today,
        ~BreakRecord.break_type.in_(['punch_in', 'punch_out'])
    ).count()
    
    # Active breaks (exclude punch_in/punch_out)
    all_active = BreakRecord.query.filter_by(end_time=None).all()
    active_breaks = len([b for b in all_active if b.break_type not in ['punch_in', 'punch_out']])
    
    overdue_breaks = BreakRecord.query.filter(
        db.func.date(BreakRecord.start_time) == today,
        BreakRecord.is_overdue == True,
        ~BreakRecord.break_type.in_(['punch_in', 'punch_out'])
    ).count()
    
    return render_template('dashboard.html',
        user=current_user,
        agents=agents,
        total_agents=len(agents),
        total_breaks_today=total_breaks_today,
        active_breaks=active_breaks,
        overdue_breaks=overdue_breaks,
        break_types=BREAK_INFO,
        break_durations=BREAK_DURATIONS,
        date_filter=date_filter,
        agent_filter=agent_filter,
        type_filter=type_filter
    )


# ==================== API ROUTES ====================

@app.route('/api/breaks', methods=['GET'])
@login_required
def get_breaks():
    """Get breaks for dashboard"""
    try:
        if not current_user.is_rtm():
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Get filter parameters
        start_date = request.args.get('start_date', get_local_time().strftime('%Y-%m-%d'))
        end_date = request.args.get('end_date', start_date)
        agent_id = request.args.get('agent_id', '')
        break_type = request.args.get('break_type', '')
        
        # Build query for regular breaks
        # IMPORTANT: Filter by shift start date, not break date
        # If filtering for Dec 30, show all records from shifts that STARTED on Dec 30
        # This includes breaks/punches that happened on Dec 31 if the shift started Dec 30
        extended_start_date = (datetime.strptime(start_date, '%Y-%m-%d').date() - timedelta(days=1)).strftime('%Y-%m-%d')
        extended_end_date = (datetime.strptime(end_date, '%Y-%m-%d').date() + timedelta(days=1)).strftime('%Y-%m-%d')
        
        # Get all shifts that START on the requested date range
        # Wrap in try/except in case database schema hasn't been updated
        try:
            shifts_in_range = Shift.query.filter(
                Shift.start_date >= start_date,
                Shift.start_date <= end_date
            ).all()
        except Exception as shift_error:
            # If start_date column doesn't exist, return empty list
            print(f"Warning: Could not query shifts by start_date: {shift_error}")
            shifts_in_range = []
        
        # Get agent IDs from shifts in range
        agent_ids_in_range = {s.agent_id for s in shifts_in_range}
        
        # Query breaks - extend range to catch overnight shifts
        query = BreakRecord.query.filter(
            db.func.date(BreakRecord.start_time) >= extended_start_date,
            db.func.date(BreakRecord.start_time) <= extended_end_date
        )
        
        if agent_id:
            query = query.filter_by(agent_id=int(agent_id))
        
        if break_type:
            query = query.filter_by(break_type=break_type)
        
        breaks = query.order_by(BreakRecord.start_time.desc()).all()
        
        # Separate attendance records (punch_in/punch_out) from breaks
        attendance_records = [br for br in breaks if br.break_type in ['punch_in', 'punch_out']]
        regular_breaks = [br for br in breaks if br.break_type not in ['punch_in', 'punch_out']]
        
        # For attendance records, also fetch punch outs that pair with punch ins in the date range
        # This handles cases where punch in is on day 1 and punch out is on day 2
        # This works for both regular breaks and manually created breaks
        if attendance_records:
            # Find punch outs that might be paired with punch ins in the date range
            # Look for punch outs within 2 days after each punch in (to handle overnight shifts)
            extended_attendance = list(attendance_records)
            
            for punch_in in attendance_records:
                if punch_in.break_type == 'punch_in' and punch_in.start_time:
                    # Look for punch out within next 2 days (to handle overnight shifts)
                    punch_in_time = punch_in.start_time
                    max_punch_out_time = punch_in_time + timedelta(days=2)
                    
                    # Query for matching punch out (same agent, after punch in, within 2 days)
                    punch_out_query = BreakRecord.query.filter(
                        BreakRecord.agent_id == punch_in.agent_id,
                        BreakRecord.break_type == 'punch_out',
                        BreakRecord.start_time > punch_in_time,
                        BreakRecord.start_time <= max_punch_out_time
                    )
                    
                    if agent_id:
                        punch_out_query = punch_out_query.filter_by(agent_id=int(agent_id))
                    
                    matching_punch_outs = punch_out_query.all()
                    
                    # Add punch outs that aren't already in the list
                    for po in matching_punch_outs:
                        if po not in extended_attendance:
                            extended_attendance.append(po)
            
            # Also find punch ins that pair with punch outs in the date range
            # (in case punch out was created first or manually added)
            for punch_out in attendance_records:
                if punch_out.break_type == 'punch_out' and punch_out.start_time:
                    # Look for punch in within 2 days before this punch out
                    punch_out_time = punch_out.start_time
                    min_punch_in_time = punch_out_time - timedelta(days=2)
                    
                    # Query for matching punch in (same agent, before punch out, within 2 days)
                    punch_in_query = BreakRecord.query.filter(
                        BreakRecord.agent_id == punch_out.agent_id,
                        BreakRecord.break_type == 'punch_in',
                        BreakRecord.start_time >= min_punch_in_time,
                        BreakRecord.start_time < punch_out_time
                    )
                    
                    if agent_id:
                        punch_in_query = punch_in_query.filter_by(agent_id=int(agent_id))
                    
                    matching_punch_ins = punch_in_query.all()
                    
                    # Add punch ins that aren't already in the list
                    for pi in matching_punch_ins:
                        if pi not in extended_attendance:
                            extended_attendance.append(pi)
            
            attendance_records = extended_attendance
        
        # Group breaks by shift period (not just calendar date)
        # For overnight shifts (e.g., 4pm-1am), breaks after midnight belong to the shift that started the previous day
        # First, get all shifts that start or end in the date range (extended range to catch overnight shifts)
        try:
            # Get shifts where start_date or end_date falls within the extended range
            # This catches shifts that overlap with the date range
            all_shifts = Shift.query.filter(
                db.or_(
                    db.and_(
                        Shift.start_date >= extended_start_date,
                        Shift.start_date <= extended_end_date
                    ),
                    db.and_(
                        Shift.end_date >= extended_start_date,
                        Shift.end_date <= extended_end_date
                    ),
                    db.and_(
                        Shift.start_date <= extended_start_date,
                        Shift.end_date >= extended_end_date
                    )
                )
            ).all()
        except Exception as e:
            # If database schema hasn't been updated yet (start_date/end_date don't exist), 
            # try to use the old shift_date column as fallback
            try:
                # Fallback: try with start_date only
                all_shifts = Shift.query.filter(
                    Shift.start_date >= extended_start_date,
                    Shift.start_date <= extended_end_date
                ).all()
            except:
                # If that also fails, return empty list (no shifts found or schema mismatch)
                print(f"Warning: Could not query shifts - schema may need migration: {e}")
                all_shifts = []
        
        # Create a mapping: agent_id -> list of shifts
        shifts_by_agent = {}
        for shift in all_shifts:
            if shift.agent_id not in shifts_by_agent:
                shifts_by_agent[shift.agent_id] = []
            shifts_by_agent[shift.agent_id].append(shift)
        
        # Function to find which shift a break belongs to (based on punch in time or shift time)
        def find_shift_for_break(break_record, agent_shifts):
            """Find the shift that a break belongs to"""
            if not break_record.start_time:
                return None
            
            break_time = break_record.start_time
            
            # First, try to find punch in for this agent before this break (within last 24 hours)
            punch_in = BreakRecord.query.filter(
                BreakRecord.agent_id == break_record.agent_id,
                BreakRecord.break_type == 'punch_in',
                BreakRecord.start_time <= break_time,
                BreakRecord.start_time >= break_time - timedelta(hours=24)
            ).order_by(BreakRecord.start_time.desc()).first()
            
            if punch_in and punch_in.start_time:
                # Find shift that matches this punch in date
                punch_in_date = punch_in.start_time.date()
                for shift in agent_shifts:
                    if shift.start_date <= punch_in_date <= shift.end_date:
                        return shift
            
            # If no punch in found, try to match by break time and shift date range
            # For overnight shifts, break might be on next day but belong to previous day's shift
            break_date = break_time.date()
            
            for shift in agent_shifts:
                # Check if break date falls within shift date range
                if shift.start_date <= break_date <= shift.end_date:
                    # Check if break time is within reasonable range (within 24 hours of shift start)
                    shift_start_datetime = datetime.combine(shift.start_date, shift.start_time)
                    time_diff = (break_time - shift_start_datetime).total_seconds() / 3600  # hours
                    if 0 <= time_diff <= 24:  # Within 24 hours of shift start
                        return shift
            
            return None
        
        # Group breaks by agent and shift period
        # KEY CHANGE: Only show breaks that belong to shifts that STARTED in the requested date range
        agents_data = {}
        for br in regular_breaks:
            # Find the shift this break belongs to
            agent_shifts = shifts_by_agent.get(br.agent_id, [])
            shift = find_shift_for_break(br, agent_shifts)
            
            # Filter: Only include if shift started in the requested date range
            if shift:
                shift_start_date_str = shift.start_date.strftime('%Y-%m-%d')
                # Only include breaks from shifts that STARTED in the date range
                if shift_start_date_str < start_date or shift_start_date_str > end_date:
                    continue  # Skip - shift didn't start in the requested date range
            else:
                # No shift found - only include if break date is in range (fallback)
                break_date = br.start_time.date() if br.start_time else None
                if break_date:
                    break_date_str = break_date.strftime('%Y-%m-%d')
                    if break_date_str < start_date or break_date_str > end_date:
                        continue  # Skip - break date outside range and no shift
            
            if br.agent_id not in agents_data:
                agents_data[br.agent_id] = {
                    'agent_name': br.agent.full_name,
                    'breaks': [],
                    'attendance': []
                }
            
            # Add shift date info to break dict for grouping
            break_dict = br.to_dict()
            if shift:
                # Use shift start date as the grouping key (even if break is on next day)
                break_dict['shift_date'] = shift.start_date.isoformat()
            else:
                # No shift found, use break's calendar date
                break_dict['shift_date'] = br.start_time.date().isoformat() if br.start_time else None
            
            agents_data[br.agent_id]['breaks'].append(break_dict)
        
        # Group attendance records by agent and pair punch in/out together
        # Show attendance records that:
        # 1. Have punch date in the date range (primary - shows punches on the day they happened), OR
        # 2. Belong to shifts that STARTED in the date range (for overnight shifts - punch out on next day)
        # IMPORTANT: Punch in on Dec 29 should show on Dec 29, not Dec 28
        filtered_attendance = []
        for br in attendance_records:
            if not br.start_time:
                continue
            
            punch_date = br.start_time.date()
            punch_date_str = punch_date.strftime('%Y-%m-%d')
            
            # Find the shift this punch belongs to
            agent_shifts = shifts_by_agent.get(br.agent_id, [])
            shift = None
            
            # Try to find shift by punch date (check if punch date falls within shift date range)
            for s in agent_shifts:
                if s.start_date <= punch_date <= s.end_date:
                    shift = s
                    break
            
            # Include if:
            # 1. Punch date is in the requested date range (primary - shows on the day it happened)
            should_include = False
            
            if punch_date_str >= start_date and punch_date_str <= end_date:
                # Punch happened in the date range - include it
                should_include = True
            elif shift:
                # Punch doesn't match date range, but check if it belongs to a shift that started in range
                # This is for overnight shifts where punch out happens on next day
                shift_start_date_str = shift.start_date.strftime('%Y-%m-%d')
                if shift_start_date_str >= start_date and shift_start_date_str <= end_date:
                    # Only include punch OUT (not punch IN) from overnight shifts
                    # Punch IN should always show on the day it happened
                    if br.break_type == 'punch_out':
                        # Punch out from overnight shift - include it
                        should_include = True
                    # Punch in should not be included if it's on a different day than the shift start
                    # (it's a new shift, should show on its own day)
            
            if should_include:
                filtered_attendance.append(br)
        
        attendance_records = filtered_attendance
        
        # Sort attendance records by time
        attendance_records.sort(key=lambda x: x.start_time if x.start_time else datetime.min)
        
        # Group by agent first
        agent_attendance = {}
        for br in attendance_records:
            if br.agent_id not in agent_attendance:
                agent_attendance[br.agent_id] = []
            agent_attendance[br.agent_id].append(br)
        
        # Pair punch in with punch out for each agent
        # IMPORTANT: Initialize agents_data for agents that only have punches (no breaks)
        for agent_id, records in agent_attendance.items():
            if agent_id not in agents_data:
                # Get agent name from first record
                agent_name = records[0].agent.full_name if records else 'Unknown'
                agents_data[agent_id] = {
                    'agent_name': agent_name,
                    'breaks': [],
                    'attendance': []
                }
            
            # Pair punch in with the next punch out
            i = 0
            while i < len(records):
                current = records[i]
                if current.break_type == 'punch_in':
                    # Look for the next punch out (could be on same day or next day)
                    punch_out = None
                    for j in range(i + 1, len(records)):
                        if records[j].break_type == 'punch_out':
                            punch_out = records[j]
                            i = j + 1  # Skip the punch out in next iteration
                            break
                    
                    # Create combined attendance record
                    agents_data[agent_id]['attendance'].append({
                        'id': current.id,
                        'type': 'punch_pair',
                        'punch_in': {
                            'id': current.id,
                            'time': current.start_time.isoformat() if current.start_time else None,
                            'screenshot': current.start_screenshot or current.end_screenshot,
                            'date': current.start_time.date().isoformat() if current.start_time else None
                        },
                        'punch_out': {
                            'id': punch_out.id if punch_out else None,
                            'time': punch_out.start_time.isoformat() if punch_out and punch_out.start_time else None,
                            'screenshot': (punch_out.start_screenshot or punch_out.end_screenshot) if punch_out else None,
                            'date': punch_out.start_time.date().isoformat() if punch_out and punch_out.start_time else None
                        } if punch_out else None,
                        'notes': current.notes or ''
                    })
                    
                    if not punch_out:
                        i += 1  # No punch out found, move to next
                else:
                    # Standalone punch out (no matching punch in) - show separately
                    agents_data[agent_id]['attendance'].append({
                        'id': current.id,
                        'type': 'punch_out',
                        'name': 'Punch Out',
                        'emoji': '🔴',
                        'time': current.start_time.isoformat() if current.start_time else None,
                        'screenshot': current.start_screenshot or current.end_screenshot,
                        'notes': current.notes or ''
                    })
                    i += 1
        
        return jsonify({
            'agents': list(agents_data.values()),
            'total_breaks': len(regular_breaks)  # Only count regular breaks
        })
    except Exception as e:
        # Log the error for debugging
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error in get_breaks: {e}")
        print(error_trace)
        # Return a proper JSON error response
        return jsonify({
            'error': f'Failed to load breaks: {str(e)}',
            'details': str(e) if DEBUG else 'Internal server error'
        }), 500


@app.route('/api/break/start', methods=['POST'])
@login_required
def start_break():
    """Start a new break"""
    if current_user.is_rtm():
        return jsonify({'error': 'RTM cannot take breaks'}), 403
    
    # Check for active break (exclude punch_in/punch_out as they're auto-completed instantly)
    all_active = BreakRecord.query.filter(
        BreakRecord.agent_id == current_user.id,
        BreakRecord.end_time == None
    ).all()
    active = next((b for b in all_active if b.break_type not in ['punch_in', 'punch_out']), None)
    if active:
        return jsonify({'error': 'You already have an active break'}), 400
    
    break_type = request.form.get('break_type')
    screenshot = request.files.get('screenshot')
    
    if not break_type or break_type not in BREAK_DURATIONS:
        return jsonify({'error': 'Invalid break type'}), 400
    
    # Require punch_in before other breaks (except punch_in itself)
    # Also prevent breaks if already punched out
    # IMPORTANT: Check if agent is still in an active shift period (overnight shifts)
    if break_type != 'punch_in':
        now = get_local_time().replace(tzinfo=None)
        today = now.date()
        
        # For punch_out: Just check if there's a punch_in that hasn't been punched out yet
        # For other breaks: Check punch_in and shift period
        if break_type == 'punch_out':
            # Find the most recent punch_in (any time, not just today)
            punch_in = BreakRecord.query.filter(
                BreakRecord.agent_id == current_user.id,
                BreakRecord.break_type == 'punch_in'
            ).order_by(BreakRecord.start_time.desc()).first()
            
            if not punch_in:
                return jsonify({
                    'error': 'You must punch in first before punching out. Please punch in to continue.'
                }), 400
            
            # Check if there's already a punch_out after this punch_in
            punch_out = BreakRecord.query.filter(
                BreakRecord.agent_id == current_user.id,
                BreakRecord.break_type == 'punch_out',
                BreakRecord.start_time > punch_in.start_time
            ).order_by(BreakRecord.start_time.desc()).first()
            
            if punch_out:
                return jsonify({
                    'error': 'You have already punched out after your last punch in.'
                }), 400
        else:
            # For regular breaks: Check for punch in today OR within last 24 hours (to handle overnight shifts)
            # First check for today's punch in
            punch_in = BreakRecord.query.filter(
                BreakRecord.agent_id == current_user.id,
                BreakRecord.break_type == 'punch_in',
                db.func.date(BreakRecord.start_time) == today
            ).first()
            
            # If no punch in today, check for punch in within last 24 hours (for overnight shifts)
            if not punch_in:
                twenty_four_hours_ago = now - timedelta(hours=24)
                punch_in = BreakRecord.query.filter(
                    BreakRecord.agent_id == current_user.id,
                    BreakRecord.break_type == 'punch_in',
                    BreakRecord.start_time >= twenty_four_hours_ago,
                    BreakRecord.start_time <= now
                ).order_by(BreakRecord.start_time.desc()).first()
            
            # IMPORTANT: If punch_in exists, allow breaks regardless of shift period
            # The shift period check was too strict - as long as there's a punch_in and no punch_out,
            # the agent should be able to take breaks
            # Shift period is only used for determining if we need to look in last 24 hours (overnight shifts)
            
            if not punch_in:
                return jsonify({
                    'error': 'You must punch in first before taking any breaks. Please punch in to continue.'
                }), 400
            
            # Check if already punched out (today or within last 24 hours, and after punch in)
            punch_out = BreakRecord.query.filter(
                BreakRecord.agent_id == current_user.id,
                BreakRecord.break_type == 'punch_out',
                db.func.date(BreakRecord.start_time) == today
            ).first()
            
            # If no punch out today, check within last 24 hours
            if not punch_out:
                twenty_four_hours_ago = now - timedelta(hours=24)
                punch_out = BreakRecord.query.filter(
                    BreakRecord.agent_id == current_user.id,
                    BreakRecord.break_type == 'punch_out',
                    BreakRecord.start_time >= twenty_four_hours_ago,
                    BreakRecord.start_time <= now
                ).order_by(BreakRecord.start_time.desc()).first()
            
            # If punched out, check if it's after the punch in
            if punch_out and punch_in:
                if punch_out.start_time > punch_in.start_time:
                    return jsonify({
                        'error': 'You have already punched out for the day. Breaks are no longer available.'
                    }), 400
    
    if not screenshot:
        return jsonify({'error': 'Screenshot is required'}), 400
    
    # Save screenshot
    screenshot_path = save_screenshot(screenshot)
    if not screenshot_path:
        return jsonify({'error': 'Invalid screenshot file'}), 400
    
    # Create break record
    break_record = BreakRecord(
        agent_id=current_user.id,
        break_type=break_type,
        start_time=get_local_time().replace(tzinfo=None),
        start_screenshot=screenshot_path
    )
    
    # Auto-complete punch_in and punch_out instantly
    if break_type in ['punch_in', 'punch_out']:
        break_record.end_time = break_record.start_time
        break_record.end_screenshot = screenshot_path
        break_record.duration_minutes = 0
        break_record.is_overdue = False
        
        # Prevent duplicate punch records for the same day
        if break_type == 'punch_in':
            today = get_local_time().date()
            existing = BreakRecord.query.filter(
                BreakRecord.agent_id == current_user.id,
                BreakRecord.break_type == 'punch_in',
                db.func.date(BreakRecord.start_time) == today
            ).first()
            if existing:
                return jsonify({'error': 'You have already punched in today'}), 400
        elif break_type == 'punch_out':
            today = get_local_time().date()
            existing = BreakRecord.query.filter(
                BreakRecord.agent_id == current_user.id,
                BreakRecord.break_type == 'punch_out',
                db.func.date(BreakRecord.start_time) == today
            ).first()
            if existing:
                return jsonify({'error': 'You have already punched out today'}), 400
    
    db.session.add(break_record)
    db.session.commit()
    
    if break_type in ['punch_in', 'punch_out']:
        action = "Punched in" if break_type == 'punch_in' else "Punched out"
        return jsonify({
            'success': True,
            'message': f'{action} successfully! ✅',
            'break': break_record.to_dict()
        })
    
    return jsonify({
        'success': True,
        'message': f'Break started! Return within {BREAK_DURATIONS[break_type]} minutes',
        'break': break_record.to_dict()
    })


@app.route('/api/break/end', methods=['POST'])
@login_required
def end_break():
    """End current break"""
    if current_user.is_rtm():
        return jsonify({'error': 'RTM cannot take breaks'}), 403
    
    # Get active break (exclude punch_in/punch_out as they're auto-completed instantly)
    all_active = BreakRecord.query.filter(
        BreakRecord.agent_id == current_user.id,
        BreakRecord.end_time == None
    ).all()
    active = next((b for b in all_active if b.break_type not in ['punch_in', 'punch_out']), None)
    if not active:
        return jsonify({'error': 'No active break to end'}), 400
    
    screenshot = request.files.get('screenshot')
    if not screenshot:
        return jsonify({'error': 'Screenshot is required'}), 400
    
    # Save screenshot
    screenshot_path = save_screenshot(screenshot)
    if not screenshot_path:
        return jsonify({'error': 'Invalid screenshot file'}), 400
    
    # Update break record
    active.end_time = get_local_time().replace(tzinfo=None)
    active.end_screenshot = screenshot_path
    active.duration_minutes = int((active.end_time - active.start_time).total_seconds() / 60)
    
    # Working time breaks (coaching/meetings) should never be marked as overdue
    # They count as working time regardless of duration
    if active.break_type in WORKING_TIME_BREAKS:
        active.is_overdue = False
    else:
        active.is_overdue = active.duration_minutes > active.get_allowed_duration()
    
    db.session.commit()
    
    status = "on time" if not active.is_overdue else "OVERDUE"
    return jsonify({
        'success': True,
        'message': f'Break ended! Duration: {active.duration_minutes} minutes ({status})',
        'break': active.to_dict()
    })


@app.route('/api/break/<int:break_id>/notes', methods=['POST'])
@login_required
def update_notes(break_id):
    """Update break notes"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    break_record = BreakRecord.query.get_or_404(break_id)
    notes = request.json.get('notes', '')
    
    break_record.notes = notes
    db.session.commit()
    
    return jsonify({'success': True, 'notes': notes})


@app.route('/api/break/manual', methods=['POST'])
@login_required
def create_manual_break():
    """Create a manual break record (RTM only)"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    agent_id = request.form.get('agent_id')
    break_type = request.form.get('break_type')
    start_date = request.form.get('start_date')
    start_time = request.form.get('start_time')
    end_date = request.form.get('end_date')
    end_time = request.form.get('end_time')
    notes = request.form.get('notes', '')
    start_screenshot_file = request.files.get('start_screenshot')
    end_screenshot_file = request.files.get('end_screenshot')
    
    # Validation - start date/time are required, end date/time are optional
    if not all([agent_id, break_type, start_date, start_time]):
        return jsonify({'error': 'Agent, break type, start date, and start time are required'}), 400
    
    if break_type not in BREAK_DURATIONS:
        return jsonify({'error': 'Invalid break type'}), 400
    
    # Verify agent exists
    agent = User.query.filter_by(id=int(agent_id), role=ROLE_AGENT).first()
    if not agent:
        return jsonify({'error': 'Agent not found'}), 404
    
    try:
        # Parse start date and time
        start_datetime = datetime.strptime(f'{start_date} {start_time}', '%Y-%m-%d %H:%M')
        
        # Parse end date and time (optional - defaults to start time if not provided)
        if end_date and end_time:
            end_datetime = datetime.strptime(f'{end_date} {end_time}', '%Y-%m-%d %H:%M')
            
            # For punch_in and punch_out, start and end times can be the same (instant actions)
            # For other break types, end time must be after start time
            if break_type not in ['punch_in', 'punch_out']:
                if end_datetime <= start_datetime:
                    return jsonify({'error': 'End time must be after start time'}), 400
            else:
                # For punch_in/punch_out, if end time is before start time, set it equal to start time
                if end_datetime < start_datetime:
                    end_datetime = start_datetime
        elif end_time:
            # End time provided but no end date - use start date
            end_datetime = datetime.strptime(f'{start_date} {end_time}', '%Y-%m-%d %H:%M')
            
            # For punch_in and punch_out, start and end times can be the same
            if break_type not in ['punch_in', 'punch_out']:
                if end_datetime <= start_datetime:
                    return jsonify({'error': 'End time must be after start time'}), 400
            else:
                if end_datetime < start_datetime:
                    end_datetime = start_datetime
        else:
            # End time not provided - use start time (for instant actions like punch_in/punch_out)
            # or leave as None for active breaks
            if break_type in ['punch_in', 'punch_out']:
                # For punch_in/punch_out, end time equals start time (instant action)
                end_datetime = start_datetime
            else:
                # For other breaks, if end time not provided, leave it as None (active break)
                end_datetime = None
        
        # Save screenshots if provided
        start_screenshot_path = None
        end_screenshot_path = None
        
        if start_screenshot_file:
            start_screenshot_path = save_screenshot(start_screenshot_file)
            if not start_screenshot_path:
                return jsonify({'error': 'Invalid start screenshot file'}), 400
        
        if end_screenshot_file:
            end_screenshot_path = save_screenshot(end_screenshot_file)
            if not end_screenshot_path:
                return jsonify({'error': 'Invalid end screenshot file'}), 400
        
        # For punch_in/punch_out, check for existing records on the same day
        # and warn if there's already one, but allow creation (RTM override)
        if break_type in ['punch_in', 'punch_out']:
            punch_date = start_datetime.date()
            existing = BreakRecord.query.filter(
                BreakRecord.agent_id == int(agent_id),
                BreakRecord.break_type == break_type,
                db.func.date(BreakRecord.start_time) == punch_date
            ).first()
            
            if existing:
                # Allow RTM to create duplicate if needed (maybe correcting a mistake)
                # But we'll still create the record
                pass
        
        # Calculate duration
        # For punch_in/punch_out, duration is 0 (instant actions)
        # For other breaks, calculate actual duration if end time is provided
        if break_type in ['punch_in', 'punch_out']:
            duration_minutes = 0
            # For punch_in/punch_out, use the same time for both start and end
            end_datetime = start_datetime
        elif end_datetime:
            duration_minutes = int((end_datetime - start_datetime).total_seconds() / 60)
        else:
            # Active break (no end time provided)
            duration_minutes = None
        
        # Create break record
        break_record = BreakRecord(
            agent_id=int(agent_id),
            break_type=break_type,
            start_time=start_datetime.replace(tzinfo=None),
            end_time=end_datetime.replace(tzinfo=None) if end_datetime else None,
            start_screenshot=start_screenshot_path,
            end_screenshot=end_screenshot_path,
            duration_minutes=duration_minutes,
            notes=notes
        )
        
        # Set is_overdue based on break type
        # punch_in/punch_out and working time breaks are never overdue
        # Active breaks (no end time) are not overdue yet
        if break_type in ['punch_in', 'punch_out'] or break_type in WORKING_TIME_BREAKS:
            break_record.is_overdue = False
        elif end_datetime and duration_minutes is not None:
            break_record.is_overdue = duration_minutes > break_record.get_allowed_duration()
        else:
            # Active break - not overdue yet
            break_record.is_overdue = False
        
        db.session.add(break_record)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Manual break created successfully for {agent.full_name}',
            'break': break_record.to_dict()
        })
        
    except ValueError as e:
        return jsonify({'error': f'Invalid date or time format: {str(e)}'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to create break: {str(e)}'}), 500


@app.route('/api/fix/working-time-breaks', methods=['POST'])
@login_required
def fix_working_time_breaks_endpoint():
    """Manual fix endpoint to clear is_overdue for working time breaks"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    fixed_count = fix_existing_working_time_breaks()
    
    return jsonify({
        'success': True,
        'message': f'Fixed {fixed_count} working time breaks',
        'fixed_count': fixed_count
    })


@app.route('/api/agents', methods=['GET'])
@login_required
def get_agents():
    """Get all agents"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    agents = User.query.filter_by(role=ROLE_AGENT).order_by(User.full_name).all()
    return jsonify({
        'agents': [{'id': a.id, 'username': a.username, 'full_name': a.full_name} for a in agents]
    })


@app.route('/api/agent/create', methods=['POST'])
@login_required
def create_agent():
    """Create a new agent"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '')
    full_name = data.get('full_name', '').strip()
    
    if not username or not password or not full_name:
        return jsonify({'error': 'All fields are required'}), 400
    
    if len(password) < 4:
        return jsonify({'error': 'Password must be at least 4 characters'}), 400
    
    if User.query.filter_by(username=username).first():
        return jsonify({'error': 'Username already exists'}), 400
    
    user = User(username=username, full_name=full_name, role=ROLE_AGENT)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'agent': {'id': user.id, 'username': user.username, 'full_name': user.full_name}
    })


@app.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    """Serve uploaded files"""
    try:
        upload_folder = Path(app.config['UPLOAD_FOLDER'])
        file_path = upload_folder / filename
        
        # Check if file exists with the given path
        if file_path.exists() and file_path.is_file():
            # File exists - serve it
            # send_from_directory needs the directory and filename separately
            if '/' in filename:
                # Nested path like "2025-12-31/abc123.jpg"
                directory = upload_folder / filename.rsplit('/', 1)[0]
                file_only = filename.rsplit('/', 1)[1]
                return send_from_directory(str(directory), file_only)
            else:
                # Just filename, no date folder
                return send_from_directory(str(upload_folder), filename)
        
        # File not found - try backwards compatibility (look in date folders)
        if '/' not in filename:
            # Try to find it in any date folder
            for date_folder in upload_folder.iterdir():
                if date_folder.is_dir():
                    potential_path = date_folder / filename
                    if potential_path.exists() and potential_path.is_file():
                        return send_from_directory(str(date_folder), filename)
        
        # File not found - return 404
        from flask import abort
        abort(404)
            
    except Exception as e:
        # Log error for debugging
        import logging
        logging.error(f"Error serving file {filename}: {str(e)}")
        from flask import abort
        abort(404)


# ==================== SHIFT MANAGEMENT API ====================

@app.route('/api/shifts', methods=['GET'])
@login_required
def get_shifts():
    """Get shifts for a date range"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    start_date = request.args.get('start_date', get_local_time().strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', start_date)
    agent_id = request.args.get('agent_id', '')
    
    query = Shift.query.filter(
        Shift.start_date >= start_date,
        Shift.start_date <= end_date
    )
    
    if agent_id:
        query = query.filter_by(agent_id=int(agent_id))
    
    shifts = query.order_by(Shift.start_date, Shift.start_time).all()
    
    return jsonify({
        'shifts': [s.to_dict() for s in shifts],
        'total': len(shifts)
    })


@app.route('/api/shift/bulk', methods=['POST'])
@login_required
def create_bulk_shifts():
    """Create shifts for multiple agents over a date range (e.g., 2 weeks)"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.json
    agent_ids = data.get('agent_ids', [])
    start_date_str = data.get('start_date')
    end_date_str = data.get('end_date')
    start_time = data.get('start_time')
    end_time = data.get('end_time')
    
    if not all([agent_ids, start_date_str, end_date_str, start_time, end_time]):
        return jsonify({'error': 'All fields are required'}), 400
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        start_time_obj = datetime.strptime(start_time, '%H:%M').time()
        end_time_obj = datetime.strptime(end_time, '%H:%M').time()
    except ValueError:
        return jsonify({'error': 'Invalid date or time format'}), 400
    
    # Validate date range (max 14 days for 2 weeks)
    days_diff = (end_date - start_date).days + 1
    if days_diff > 14:
        return jsonify({'error': 'Date range cannot exceed 14 days (2 weeks)'}), 400
    
    if days_diff < 1:
        return jsonify({'error': 'End date must be after start date'}), 400
    
    created = 0
    updated = 0
    
    # Generate all dates in range
    current_date = start_date
    dates_to_create = []
    while current_date <= end_date:
        dates_to_create.append(current_date)
        current_date += timedelta(days=1)
    
    # Create/update shifts for each agent and each date
    for agent_id in agent_ids:
        agent = User.query.filter_by(id=agent_id, role=ROLE_AGENT).first()
        if not agent:
            continue
        
        for shift_date in dates_to_create:
            # For bulk creation, end_date is same as start_date (single day shifts)
            existing = Shift.query.filter_by(
                agent_id=agent_id,
                start_date=shift_date,
                end_date=shift_date
            ).first()
            if existing:
                existing.start_time = start_time_obj
                existing.end_time = end_time_obj
                updated += 1
            else:
                shift = Shift(
                    agent_id=agent_id,
                    start_date=shift_date,
                    start_time=start_time_obj,
                    end_date=shift_date,
                    end_time=end_time_obj,
                    created_by=current_user.id
                )
                db.session.add(shift)
                created += 1
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f'Created {created} shifts, updated {updated} shifts for {len(dates_to_create)} days',
        'days_created': len(dates_to_create),
        'created': created,
        'updated': updated
    })


@app.route('/api/shift/<int:shift_id>', methods=['DELETE'])
@login_required
def delete_shift(shift_id):
    """Delete a shift"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    shift = Shift.query.get_or_404(shift_id)
    db.session.delete(shift)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Shift deleted'})


@app.route('/api/shift', methods=['POST'])
@login_required
def create_shift():
    """Create a single shift for an agent"""
    try:
        if not current_user.is_rtm():
            return jsonify({'error': 'Unauthorized'}), 403
        
        if not request.json:
            return jsonify({'error': 'Request body must be JSON'}), 400
        
        data = request.json
        agent_id = data.get('agent_id')
        start_date_str = data.get('start_date')
        start_time_str = data.get('start_time')
        end_date_str = data.get('end_date')
        end_time_str = data.get('end_time')
        
        if not all([agent_id, start_date_str, start_time_str, end_date_str, end_time_str]):
            return jsonify({'error': 'All fields are required'}), 400
        
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            start_time = datetime.strptime(start_time_str, '%H:%M').time()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            end_time = datetime.strptime(end_time_str, '%H:%M').time()
        except ValueError as e:
            return jsonify({'error': f'Invalid date or time format: {str(e)}'}), 400
        
        # Validate that end date/time is after start date/time
        start_datetime = datetime.combine(start_date, start_time)
        end_datetime = datetime.combine(end_date, end_time)
        if end_datetime <= start_datetime:
            return jsonify({'error': 'End date/time must be after start date/time'}), 400
        
        # Check if shift already exists for this agent with same start date/time
        existing = Shift.query.filter_by(
            agent_id=agent_id,
            start_date=start_date,
            start_time=start_time
        ).first()
        
        if existing:
            # Update existing shift
            existing.end_date = end_date
            existing.end_time = end_time
            existing.created_by = current_user.id
            db.session.commit()
            return jsonify({
                'success': True,
                'message': 'Shift updated',
                'shift': existing.to_dict()
            })
        
        # Create new shift
        shift = Shift(
            agent_id=agent_id,
            start_date=start_date,
            start_time=start_time,
            end_date=end_date,
            end_time=end_time,
            created_by=current_user.id
        )
        db.session.add(shift)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Shift created',
            'shift': shift.to_dict()
        })
    except Exception as e:
        # Log the error for debugging
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error in create_shift: {e}")
        print(error_trace)
        # Return a proper JSON error response
        return jsonify({
            'error': f'Failed to create shift: {str(e)}',
            'details': str(e) if DEBUG else 'Internal server error'
        }), 500


@app.route('/api/shift/<int:shift_id>', methods=['PUT'])
@login_required
def update_shift(shift_id):
    """Update an existing shift"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    shift = Shift.query.get_or_404(shift_id)
    data = request.json
    
    if 'start_date' in data:
        shift.start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
    if 'start_time' in data:
        shift.start_time = datetime.strptime(data['start_time'], '%H:%M').time()
    if 'end_date' in data:
        shift.end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
    if 'end_time' in data:
        shift.end_time = datetime.strptime(data['end_time'], '%H:%M').time()
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Shift updated',
        'shift': shift.to_dict()
    })


# ==================== OFF DAYS MANAGEMENT ====================

@app.route('/api/offdays', methods=['GET'])
@login_required
def get_offdays():
    """Get off days for agents"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    agent_id = request.args.get('agent_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = OffDay.query
    
    if agent_id:
        query = query.filter_by(agent_id=agent_id)
    
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(OffDay.off_date >= start, OffDay.off_date <= end)
        except ValueError:
            return jsonify({'error': 'Invalid date format'}), 400
    
    offdays = query.order_by(OffDay.off_date.desc()).all()
    
    return jsonify({
        'offdays': [od.to_dict() for od in offdays],
        'total': len(offdays)
    })


@app.route('/api/offday', methods=['POST'])
@login_required
def create_offday():
    """Create an off day for an agent"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.json
    agent_id = data.get('agent_id')
    off_date_str = data.get('off_date')
    reason = data.get('reason', '')
    
    if not all([agent_id, off_date_str]):
        return jsonify({'error': 'Agent ID and date are required'}), 400
    
    try:
        off_date = datetime.strptime(off_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400
    
    # Check if off day already exists
    existing = OffDay.query.filter_by(
        agent_id=agent_id,
        off_date=off_date
    ).first()
    
    if existing:
        # Update existing off day
        existing.reason = reason
        existing.created_by = current_user.id
        db.session.commit()
        return jsonify({
            'success': True,
            'message': 'Off day updated',
            'offday': existing.to_dict()
        })
    
    # Create new off day
    offday = OffDay(
        agent_id=agent_id,
        off_date=off_date,
        reason=reason,
        created_by=current_user.id
    )
    db.session.add(offday)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Off day created',
        'offday': offday.to_dict()
    })


@app.route('/api/offday/<int:offday_id>', methods=['PUT'])
@login_required
def update_offday(offday_id):
    """Update an existing off day"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    offday = OffDay.query.get_or_404(offday_id)
    data = request.json
    
    if 'off_date' in data:
        offday.off_date = datetime.strptime(data['off_date'], '%Y-%m-%d').date()
    if 'reason' in data:
        offday.reason = data['reason']
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Off day updated',
        'offday': offday.to_dict()
    })


@app.route('/api/offday/<int:offday_id>', methods=['DELETE'])
@login_required
def delete_offday(offday_id):
    """Delete an off day"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    offday = OffDay.query.get_or_404(offday_id)
    db.session.delete(offday)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Off day deleted'})


# ==================== REPORTING & EXPORT ====================

def calculate_agent_metrics(agent_id, start_date, end_date):
    """Calculate all metrics for an agent within a date range"""
    
    # Get all breaks for this agent in date range
    breaks = BreakRecord.query.filter(
        BreakRecord.agent_id == agent_id,
        db.func.date(BreakRecord.start_time) >= start_date,
        db.func.date(BreakRecord.start_time) <= end_date
    ).all()
    
    # Get all shifts for this agent in date range (shifts that start in the range)
    shifts = Shift.query.filter(
        Shift.agent_id == agent_id,
        Shift.start_date >= datetime.strptime(start_date, '%Y-%m-%d').date(),
        Shift.start_date <= datetime.strptime(end_date, '%Y-%m-%d').date()
    ).all()
    
    # Calculate metrics
    total_scheduled_minutes = sum(s.get_duration_hours() * 60 for s in shifts)
    
    # Separate breaks into working time breaks and regular breaks
    working_time_breaks = [b for b in breaks if b.break_type in WORKING_TIME_BREAKS and b.end_time]
    # Regular breaks include emergency (emergency counts as break time, not working time)
    regular_breaks = [b for b in breaks if b.break_type not in WORKING_TIME_BREAKS + ['punch_in', 'punch_out'] and b.end_time]
    
    # Total break minutes (including emergency - emergency counts as break time)
    # Excluding working time breaks and punch records
    total_break_minutes = sum(b.duration_minutes or 0 for b in regular_breaks)
    total_allowed_break_minutes = sum(b.get_allowed_duration() for b in regular_breaks)
    exceeding_break_minutes = max(0, total_break_minutes - total_allowed_break_minutes)
    
    # Count incidents (overdue breaks) - only for regular breaks
    # Use effective overdue status (working time breaks are never overdue)
    incidents = sum(1 for b in regular_breaks if b.get_effective_overdue_status())
    
    # Count emergency breaks
    emergency_count = sum(1 for b in breaks if b.break_type == 'emergency')
    
    # Count overtime breaks (in minutes)
    overtime_breaks = [b for b in breaks if b.break_type == 'overtime' and b.end_time]
    overtime_count = len(overtime_breaks)
    overtime_minutes = sum(b.duration_minutes or 0 for b in overtime_breaks)
    
    # Count breaks by type (exclude punch_in/punch_out as they're attendance, not breaks)
    break_counts = {}
    for b in breaks:
        if b.break_type not in ['punch_in', 'punch_out']:
            break_counts[b.break_type] = break_counts.get(b.break_type, 0) + 1
    
    # Calculate utilization
    # Working time breaks (coaching/meetings/overtime) count as working time, not breaks
    # Emergency breaks count as break time and reduce utilization
    # Expected working hours: 8 hours per shift = 480 minutes
    # Allocated break time: 1 hour 15 minutes per shift = 75 minutes
    if len(shifts) > 0:
        expected_working_minutes = len(shifts) * 8 * 60  # 8 hours per shift = 480 minutes
        expected_break_minutes = len(shifts) * 75  # 75 minutes per shift (1 hour 15 minutes allocated break time)
        
        # If breaks exceed allocated time, utilization decreases
        # Emergency breaks are included in total_break_minutes (they count as break time)
        excess_break_minutes = max(0, total_break_minutes - expected_break_minutes)
        actual_working_time = expected_working_minutes - excess_break_minutes
        utilization = (actual_working_time / expected_working_minutes) * 100
        utilization = min(100.0, max(0.0, utilization))  # Cap between 0% and 100%
    else:
        utilization = 0
    
    # Calculate adherence based on:
    # 1. Break durations (actual vs allowed)
    # 2. Punch in time vs shift start time
    # 3. Punch out time vs shift end time
    
    adherence_scores = []
    
    # 1. Break duration adherence (only for regular breaks)
    total_completed_breaks = len([b for b in regular_breaks if b.end_time])
    
    if total_completed_breaks > 0:
        for b in regular_breaks:
            if b.end_time and b.duration_minutes is not None:
                allowed_duration = b.get_allowed_duration()
                if allowed_duration > 0:
                    actual_duration = b.duration_minutes
                    if actual_duration <= allowed_duration:
                        break_adherence = 100.0
                    else:
                        break_adherence = (allowed_duration / actual_duration) * 100
                    adherence_scores.append(break_adherence)
    
    # 2. Punch in/out adherence based on shift times
    # Group shifts by start date for easier lookup
    shifts_by_date = {s.start_date: s for s in shifts}
    
    # Group punch in/out by date
    punch_records_by_date = {}
    for b in breaks:
        if b.break_type in ['punch_in', 'punch_out'] and b.start_time:
            punch_date = b.start_time.date()
            if punch_date not in punch_records_by_date:
                punch_records_by_date[punch_date] = {}
            punch_records_by_date[punch_date][b.break_type] = b
    
    # Calculate punch in/out adherence for each day with a shift
    for shift_start_date, shift in shifts_by_date.items():
        punch_records = punch_records_by_date.get(shift_start_date, {})
        
        # Punch in adherence
        if 'punch_in' in punch_records:
            punch_in = punch_records['punch_in']
            punch_in_time = punch_in.start_time.time()
            shift_start_time = shift.start_time
            
            # Calculate difference in minutes
            punch_in_datetime = datetime.combine(shift_start_date, punch_in_time)
            shift_start_datetime = datetime.combine(shift.start_date, shift_start_time)
            
            # Allow 5 minutes grace period (early or late)
            time_diff_minutes = abs((punch_in_datetime - shift_start_datetime).total_seconds() / 60)
            
            if time_diff_minutes <= 5:
                punch_in_adherence = 100.0
            else:
                # Penalty: decrease adherence for being late/early
                # Max penalty at 30 minutes = 0% adherence
                # Formula: max(0, (30 - time_diff) / 30 * 100)
                punch_in_adherence = max(0, (30 - time_diff_minutes) / 30 * 100)
            adherence_scores.append(punch_in_adherence)
        else:
            # No punch in = 0% adherence for that day
            adherence_scores.append(0.0)
        
        # Punch out adherence
        if 'punch_out' in punch_records:
            punch_out = punch_records['punch_out']
            punch_out_time = punch_out.start_time.time()
            shift_end_time = shift.end_time
            
            # Calculate difference in minutes
            punch_out_datetime = datetime.combine(punch_out.start_time.date(), punch_out_time)
            shift_end_datetime = datetime.combine(shift.end_date, shift_end_time)
            
            # Allow 5 minutes grace period (early or late)
            time_diff_minutes = abs((punch_out_datetime - shift_end_datetime).total_seconds() / 60)
            
            if time_diff_minutes <= 5:
                punch_out_adherence = 100.0
            else:
                # Penalty: decrease adherence for being late/early
                # Max penalty at 30 minutes = 0% adherence
                punch_out_adherence = max(0, (30 - time_diff_minutes) / 30 * 100)
            adherence_scores.append(punch_out_adherence)
        else:
            # No punch out = 0% adherence for that day
            adherence_scores.append(0.0)
    
    # Calculate overall adherence as average of all scores
    if adherence_scores:
        adherence = sum(adherence_scores) / len(adherence_scores)
    else:
        adherence = 100  # No data = 100% adherence (default)
    
    # Conformance: measures if agent worked the expected hours
    # Expected: 8 hours per shift
    # Actual: scheduled time - excess break time (breaks beyond allocated 75 minutes)
    # Compensation can add working time back
    if len(shifts) > 0:
        expected_working_minutes = len(shifts) * 8 * 60  # 8 hours per shift
        expected_break_minutes = len(shifts) * 75  # 75 minutes allocated break time per shift
        
        # Get compensation minutes (compensation adds working time)
        compensation_minutes = sum(
            b.duration_minutes or 0 
            for b in breaks 
            if b.break_type == 'compensation' and b.end_time
        )
        
        # Calculate actual working time
        # Emergency breaks are included in total_break_minutes (they reduce working time)
        excess_break_minutes = max(0, total_break_minutes - expected_break_minutes)
        actual_working_minutes = expected_working_minutes - excess_break_minutes + compensation_minutes
        
        conformance = min(100.0, (actual_working_minutes / expected_working_minutes) * 100)
    else:
        conformance = 0  # No shifts assigned = 0% conformance
    
    return {
        'total_scheduled_hours': round(total_scheduled_minutes / 60, 2),
        'total_break_minutes': total_break_minutes,
        'total_allowed_break_minutes': total_allowed_break_minutes,
        'exceeding_break_minutes': exceeding_break_minutes,
        'incidents': incidents,
        'emergency_count': emergency_count,
        'overtime_count': overtime_count,
        'overtime_minutes': overtime_minutes,
        'total_breaks': len(breaks),
        'completed_breaks': total_completed_breaks,
        'utilization': round(utilization, 1),
        'adherence': round(adherence, 1),
        'conformance': round(conformance, 1),
        'break_counts': break_counts,
        'shifts_count': len(shifts)
    }


@app.route('/api/report/metrics', methods=['GET'])
@login_required
def get_metrics():
    """Get metrics for all agents"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    start_date = request.args.get('start_date', get_local_time().strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', start_date)
    
    agents = User.query.filter_by(role=ROLE_AGENT).order_by(User.full_name).all()
    
    results = []
    totals = {
        'total_scheduled_hours': 0,
        'total_break_minutes': 0,
        'exceeding_break_minutes': 0,
        'incidents': 0,
        'emergency_count': 0,
        'overtime_count': 0,
        'overtime_minutes': 0,
        'total_breaks': 0,
        'utilization_sum': 0,
        'adherence_sum': 0,
        'conformance_sum': 0,
        'agent_count': 0
    }
    
    for agent in agents:
        metrics = calculate_agent_metrics(agent.id, start_date, end_date)
        results.append({
            'agent_id': agent.id,
            'agent_name': agent.full_name,
            'username': agent.username,
            **metrics
        })
        
        # Accumulate totals
        totals['total_scheduled_hours'] += metrics['total_scheduled_hours']
        totals['total_break_minutes'] += metrics['total_break_minutes']
        totals['exceeding_break_minutes'] += metrics['exceeding_break_minutes']
        totals['incidents'] += metrics['incidents']
        totals['emergency_count'] += metrics['emergency_count']
        totals['overtime_count'] += metrics['overtime_count']
        totals['overtime_minutes'] += metrics['overtime_minutes']
        totals['total_breaks'] += metrics['total_breaks']
        if metrics['shifts_count'] > 0:
            totals['utilization_sum'] += metrics['utilization']
            totals['adherence_sum'] += metrics['adherence']
            totals['conformance_sum'] += metrics['conformance']
            totals['agent_count'] += 1
    
    # Calculate averages
    if totals['agent_count'] > 0:
        totals['avg_utilization'] = round(totals['utilization_sum'] / totals['agent_count'], 1)
        totals['avg_adherence'] = round(totals['adherence_sum'] / totals['agent_count'], 1)
        totals['avg_conformance'] = round(totals['conformance_sum'] / totals['agent_count'], 1)
    else:
        totals['avg_utilization'] = 0
        totals['avg_adherence'] = 0
        totals['avg_conformance'] = 0
    
    return jsonify({
        'agents': results,
        'totals': totals,
        'date_range': {'start': start_date, 'end': end_date}
    })


@app.route('/api/report/export', methods=['GET'])
@login_required
def export_report():
    """Export metrics to Excel"""
    if not current_user.is_rtm():
        return jsonify({'error': 'Unauthorized'}), 403
    
    start_date = request.args.get('start_date', get_local_time().strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', start_date)
    
    agents = User.query.filter_by(role=ROLE_AGENT).order_by(User.full_name).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Agent Metrics"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    good_fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
    warning_fill = PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid")
    bad_fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:O1')
    ws['A1'] = f"RTA Agent Metrics Report ({start_date} to {end_date})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Headers (row 3)
    headers = [
        "Agent Name",
        "Username", 
        "Scheduled Hours",
        "Total Breaks",
        "Break Time (min)",
        "Allowed Break (min)",
        "Exceeding (min)",
        "Incidents",
        "Emergency",
        "Overtime",
        "Overtime (min)",
        "Utilization %",
        "Adherence %",
        "Conformance %",
        "Status"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    row = 4
    total_metrics = {
        'scheduled_hours': 0,
        'break_minutes': 0,
        'allowed_minutes': 0,
        'exceeding': 0,
        'incidents': 0,
        'emergency': 0,
        'breaks': 0,
        'util_sum': 0,
        'adh_sum': 0,
        'conf_sum': 0,
        'count': 0
    }
    
    for agent in agents:
        metrics = calculate_agent_metrics(agent.id, start_date, end_date)
        
        # Determine status
        if metrics['incidents'] == 0 and metrics['exceeding_break_minutes'] == 0:
            status = "✅ Good"
            status_fill = good_fill
        elif metrics['incidents'] <= 2 or metrics['exceeding_break_minutes'] <= 15:
            status = "⚠️ Warning"
            status_fill = warning_fill
        else:
            status = "❌ Needs Review"
            status_fill = bad_fill
        
        row_data = [
            agent.full_name,
            agent.username,
            metrics['total_scheduled_hours'],
            metrics['total_breaks'],
            metrics['total_break_minutes'],
            metrics['total_allowed_break_minutes'],
            metrics['exceeding_break_minutes'],
            metrics['incidents'],
            metrics['emergency_count'],
            metrics['overtime_count'],
            metrics['overtime_minutes'],
            metrics['utilization'],
            metrics['adherence'],
            metrics['conformance'],
            status
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = border
            if col == 15:  # Status column
                cell.fill = status_fill
        
        # Accumulate totals
        total_metrics['scheduled_hours'] += metrics['total_scheduled_hours']
        total_metrics['break_minutes'] += metrics['total_break_minutes']
        total_metrics['allowed_minutes'] += metrics['total_allowed_break_minutes']
        total_metrics['exceeding'] += metrics['exceeding_break_minutes']
        total_metrics['incidents'] += metrics['incidents']
        total_metrics['emergency'] += metrics['emergency_count']
        total_metrics['overtime_count'] = total_metrics.get('overtime_count', 0) + metrics['overtime_count']
        total_metrics['overtime_minutes'] = total_metrics.get('overtime_minutes', 0) + metrics['overtime_minutes']
        total_metrics['breaks'] += metrics['total_breaks']
        if metrics['shifts_count'] > 0:
            total_metrics['util_sum'] += metrics['utilization']
            total_metrics['adh_sum'] += metrics['adherence']
            total_metrics['conf_sum'] += metrics['conformance']
            total_metrics['count'] += 1
        
        row += 1
    
    # Totals/Average row
    row += 1
    total_fill = PatternFill(start_color="e0e0e0", end_color="e0e0e0", fill_type="solid")
    
    avg_util = round(total_metrics['util_sum'] / total_metrics['count'], 1) if total_metrics['count'] > 0 else 0
    avg_adh = round(total_metrics['adh_sum'] / total_metrics['count'], 1) if total_metrics['count'] > 0 else 0
    avg_conf = round(total_metrics['conf_sum'] / total_metrics['count'], 1) if total_metrics['count'] > 0 else 0
    
    totals_row = [
        "TOTAL / AVERAGE",
        f"{len(agents)} agents",
        total_metrics['scheduled_hours'],
        total_metrics['breaks'],
        total_metrics['break_minutes'],
        total_metrics['allowed_minutes'],
        total_metrics['exceeding'],
        total_metrics['incidents'],
        total_metrics['emergency'],
        total_metrics.get('overtime_count', 0),
        total_metrics.get('overtime_minutes', 0),
        avg_util,
        avg_adh,
        avg_conf,
        ""
    ]
    
    for col, value in enumerate(totals_row, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=True)
        cell.alignment = cell_alignment
        cell.border = border
        cell.fill = total_fill
    
    # Adjust column widths
    column_widths = [20, 15, 15, 12, 15, 15, 12, 10, 10, 10, 12, 12, 12, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Add a summary section
    row += 3
    ws.cell(row=row, column=1, value="Summary").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value=f"Report Period: {start_date} to {end_date}")
    row += 1
    ws.cell(row=row, column=1, value=f"Total Agents: {len(agents)}")
    row += 1
    ws.cell(row=row, column=1, value=f"Total Incidents: {total_metrics['incidents']}")
    row += 1
    ws.cell(row=row, column=1, value=f"Total Emergency Breaks: {total_metrics['emergency']}")
    row += 1
    ws.cell(row=row, column=1, value=f"Total Exceeding Break Time: {total_metrics['exceeding']} minutes")
    row += 1
    ws.cell(row=row, column=1, value=f"Average Utilization: {avg_util}%")
    row += 1
    ws.cell(row=row, column=1, value=f"Average Adherence: {avg_adh}%")
    row += 1
    ws.cell(row=row, column=1, value=f"Average Conformance: {avg_conf}%")
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"RTA_Metrics_{start_date}_to_{end_date}.xlsx"
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


# ==================== STARTUP ====================

def fix_existing_working_time_breaks():
    """One-time fix: Clear is_overdue flag for existing working time breaks"""
    try:
        # Check if BreakRecord table exists by trying to query
        try:
            test_query = BreakRecord.query.limit(1).all()
        except Exception:
            print("⚠️ Database tables not initialized yet, skipping fix")
            return 0
        
        working_time_breaks = BreakRecord.query.filter(
            BreakRecord.break_type.in_(WORKING_TIME_BREAKS),
            BreakRecord.is_overdue == True
        ).all()
        
        if working_time_breaks:
            for br in working_time_breaks:
                br.is_overdue = False
            db.session.commit()
            print(f"✅ Fixed {len(working_time_breaks)} working time breaks that were incorrectly marked as overdue")
        return len(working_time_breaks)
    except Exception as e:
        print(f"⚠️ Error fixing working time breaks: {e}")
        import traceback
        traceback.print_exc()
        return 0

def create_app():
    """Initialize database - called on startup"""
    with app.app_context():
        init_db()
        # Fix any existing working time breaks that were incorrectly marked as overdue
        fix_existing_working_time_breaks()
    return app

# Initialize on import (for gunicorn)
create_app()

# ==================== MAIN ====================

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    
    print("=" * 50)
    print("RTA Break Tracker - Web Application")
    print(f"Environment: {ENV}")
    print("=" * 50)
    print(f"Open in browser: http://localhost:{port}")
    print("-" * 50)
    print("Default admin: admin / admin123")
    print("(Change password after first login!)")
    print("=" * 50)
    
    app.run(debug=DEBUG, host='0.0.0.0', port=port)

